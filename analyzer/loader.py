from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def _clean_column_name(name: Any) -> str:
    raw = "" if name is None else str(name)
    cleaned = " ".join(raw.replace("\n", " ").replace("\r", " ").split()).strip()
    return cleaned or "column"


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for col in columns:
        count = seen.get(col, 0) + 1
        seen[col] = count
        out.append(col if count == 1 else f"{col}_{count}")
    return out


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized = [_clean_column_name(col) for col in df.columns.tolist()]
    df.columns = _dedupe_columns(normalized)
    return df


def _detect_csv_layout(path: Path, encoding: str, sample_lines: int = 30) -> tuple[int, str]:
    with path.open("r", encoding=encoding, errors="strict", newline="") as f:
        lines: list[str] = []
        for _ in range(sample_lines):
            line = f.readline()
            if not line:
                break
            lines.append(line)

    text = "".join(lines) if lines else ""
    try:
        delimiter = csv.Sniffer().sniff(text, delimiters=[",", "\t", ";", "|"]).delimiter if text else ","
    except csv.Error:
        delimiter = ","

    parsed = list(csv.reader(lines, delimiter=delimiter))
    if not parsed:
        return 0, delimiter

    best_idx = 0
    best_score = -1.0
    for idx, row in enumerate(parsed[: min(15, len(parsed))]):
        cells = [c.strip() for c in row]
        if not cells:
            continue
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            continue
        unique_ratio = len(set(non_empty)) / max(1, len(non_empty))

        next_data_like = 0.0
        if idx + 1 < len(parsed):
            next_row = [c.strip() for c in parsed[idx + 1]]
            if next_row:
                non_blank_next = [c for c in next_row if c]
                data_like = sum(1 for c in non_blank_next if any(ch.isdigit() for ch in c))
                next_data_like = data_like / max(1, len(non_blank_next))

        score = (len(non_empty) * 1.0) + (unique_ratio * 2.0) + (next_data_like * 1.5)
        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx, delimiter


def load_csv(path: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    p = Path(path)
    attempts = ["utf-8-sig", "utf-8", "cp949", "euc-kr"]
    errors: list[str] = []

    for enc in attempts:
        try:
            header_idx, delimiter = _detect_csv_layout(p, encoding=enc)
            df = pd.read_csv(
                p,
                encoding=enc,
                sep=delimiter,
                skiprows=header_idx,
                header=0,
                engine="python",
            )
            df = _normalize_columns(df)
            meta = {
                "encoding": enc,
                "skiprows": header_idx,
                "sheet_name": None,
                "notes": "자동 로드 성공",
            }
            return df, meta
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{enc}: {exc}")

    raise ValueError(
        "CSV 자동 로드 실패. 인코딩/헤더를 여러 방식으로 재시도했지만 읽지 못했습니다. "
        "파일을 UTF-8로 다시 저장하거나 상단 설명행을 제거해 재시도해 주세요. "
        f"시도 내역: {' | '.join(errors)}"
    )


def _sheet_score(df: pd.DataFrame) -> int:
    non_empty_rows = df.dropna(how="all")
    if non_empty_rows.empty:
        return 0
    non_empty_cols = non_empty_rows.dropna(axis=1, how="all")
    return int(non_empty_rows.shape[0] * max(1, non_empty_cols.shape[1]))


def _formula_cache_hint(path: Path, sheet_name: str) -> str | None:
    try:
        wb_formula = load_workbook(path, data_only=False, read_only=True)
        wb_data = load_workbook(path, data_only=True, read_only=True)
        ws_formula = wb_formula[sheet_name]
        ws_data = wb_data[sheet_name]

        missing_cache = False
        for r_formula, r_data in zip(ws_formula.iter_rows(), ws_data.iter_rows()):
            for c_formula, c_data in zip(r_formula, r_data):
                if isinstance(c_formula.value, str) and c_formula.value.startswith("=") and c_data.value is None:
                    missing_cache = True
                    break
            if missing_cache:
                break

        wb_formula.close()
        wb_data.close()

        if missing_cache:
            return "수식 캐시 값이 비어 있어 일부 값이 누락될 수 있습니다. 엑셀에서 파일을 열어 저장 후 다시 시도해 주세요."
    except Exception:  # noqa: BLE001
        return None

    return None


def load_xlsx(path: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    p = Path(path)
    xls = pd.ExcelFile(p, engine="openpyxl")

    best_sheet: str | None = None
    best_df: pd.DataFrame | None = None
    best_score = -1

    for sheet in xls.sheet_names:
        df = pd.read_excel(p, sheet_name=sheet, engine="openpyxl")
        score = _sheet_score(df)
        if score > best_score:
            best_score = score
            best_sheet = sheet
            best_df = df

    if best_df is None or best_sheet is None:
        raise ValueError("XLSX에서 읽을 수 있는 시트를 찾지 못했습니다.")

    best_df = _normalize_columns(best_df)
    note = _formula_cache_hint(p, best_sheet) or "자동 로드 성공"

    meta = {
        "encoding": None,
        "skiprows": 0,
        "sheet_name": best_sheet,
        "notes": note,
    }
    return best_df, meta
